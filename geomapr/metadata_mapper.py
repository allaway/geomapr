"""
AI-powered metadata mapper that fills in CSV templates using controlled vocabularies
from the NF metadata dictionary.
"""

import json
import csv
import requests
import yaml
import os
import time
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


@dataclass
class MappingSuggestion:
    """Represents a mapping suggestion from AI"""
    original_value: str
    suggested_term: str
    confidence: float
    reasoning: str
    is_valid: bool = False  # Will be set during validation


@dataclass
class CellMapping:
    """Represents a mapping applied to a specific cell"""
    row_idx: int
    column_name: str
    mapped_value: str
    confidence: float
    reasoning: str


@dataclass
class APICallLog:
    """Represents a logged API call for tracking"""
    timestamp: datetime
    function_name: str
    model: str
    prompt_tokens: int
    completion_tokens: int
    total_tokens: int
    request_data: dict
    response_data: dict
    duration_seconds: float


class AILogger:
    """Handles logging and token tracking for AI API calls"""
    
    def __init__(self, log_file: str = "ai_api_calls.log"):
        self.log_file = log_file
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_calls = 0
        self.call_logs: List[APICallLog] = []
    
    def log_api_call(self, function_name: str, model: str, request_data: dict, 
                     response_data: dict, duration: float) -> APICallLog:
        """Log an API call and extract token usage"""
        
        # Extract token usage from response (OpenRouter Usage Accounting format)
        usage = response_data.get('usage', {})
        prompt_tokens = usage.get('prompt_tokens', 0)
        completion_tokens = usage.get('completion_tokens', 0)
        total_tokens = usage.get('total_tokens', prompt_tokens + completion_tokens)
        
        # Extract accurate cost from OpenRouter Usage Accounting
        actual_cost = usage.get('cost', 0)  # Cost in credits
        
        # Update totals
        self.total_input_tokens += prompt_tokens
        self.total_output_tokens += completion_tokens
        self.total_calls += 1
        
        # Track actual costs instead of estimated
        if not hasattr(self, 'total_actual_cost'):
            self.total_actual_cost = 0
        self.total_actual_cost += actual_cost
        
        # Create log entry
        log_entry = APICallLog(
            timestamp=datetime.now(),
            function_name=function_name,
            model=model,
            prompt_tokens=prompt_tokens,
            completion_tokens=completion_tokens,
            total_tokens=total_tokens,
            request_data=request_data,
            response_data=response_data,
            duration_seconds=duration
        )
        
        self.call_logs.append(log_entry)
        
        # Write to log file
        self._write_log_entry(log_entry, actual_cost)
        
        return log_entry
    
    def _write_log_entry(self, log_entry: APICallLog, actual_cost: float = 0):
        """Write a single log entry to file"""
        log_data = {
            'timestamp': log_entry.timestamp.isoformat(),
            'function': log_entry.function_name,
            'model': log_entry.model,
            'tokens': {
                'input': log_entry.prompt_tokens,
                'output': log_entry.completion_tokens,
                'total': log_entry.total_tokens
            },
            'cost': actual_cost,  # Actual cost from OpenRouter Usage Accounting
            'duration_seconds': log_entry.duration_seconds,
            'request': {
                'messages': log_entry.request_data.get('messages', []),
                'temperature': log_entry.request_data.get('temperature'),
                'max_tokens': log_entry.request_data.get('max_tokens')
            },
            'response_content': log_entry.response_data.get('choices', [{}])[0].get('message', {}).get('content', '')[:500] + '...' if len(log_entry.response_data.get('choices', [{}])[0].get('message', {}).get('content', '')) > 500 else log_entry.response_data.get('choices', [{}])[0].get('message', {}).get('content', '')
        }
        
        with open(self.log_file, 'a', encoding='utf-8') as f:
            f.write(json.dumps(log_data) + '\n')
    
    def print_summary(self):
        """Print a summary of token usage and costs"""
        print(f"\n{'='*60}")
        print(f"AI API USAGE SUMMARY")
        print(f"{'='*60}")
        print(f"Total API calls: {self.total_calls}")
        print(f"Total input tokens: {self.total_input_tokens:,}")
        print(f"Total output tokens: {self.total_output_tokens:,}")
        print(f"Total tokens: {(self.total_input_tokens + self.total_output_tokens):,}")
        print(f"Log file: {self.log_file}")
        
        # Show actual cost from OpenRouter Usage Accounting
        if hasattr(self, 'total_actual_cost') and self.total_actual_cost > 0:
            print(f"Actual cost: ${self.total_actual_cost:.6f} (from OpenRouter)")
        else:
            # Fallback to estimation if Usage Accounting not available
            estimated_cost = (self.total_input_tokens * 0.000001) + (self.total_output_tokens * 0.000002)
            print(f"Estimated cost: ${estimated_cost:.6f} (approximate - enable Usage Accounting for accuracy)")
        
        print(f"{'='*60}")
    
    def get_summary_dict(self) -> dict:
        """Get usage summary as dictionary"""
        return {
            'total_calls': self.total_calls,
            'total_input_tokens': self.total_input_tokens,
            'total_output_tokens': self.total_output_tokens,
            'total_tokens': self.total_input_tokens + self.total_output_tokens,
            'log_file': self.log_file
        }


@dataclass
class ColumnMapping:
    """Represents mapping configuration for a column"""
    column_name: str
    controlled_vocab: List[str]  # Valid terms for this column
    suggestions: List[MappingSuggestion]
    selected_mapping: Optional[str] = None


class NFMetadataDictionary:
    """Handles fetching and parsing of NF metadata dictionary"""
    
    def __init__(self, jsonld_url: str, cache_file: str, cache_expiry_hours: int = 24):
        self.jsonld_url = jsonld_url
        self.cache_file = cache_file
        self.cache_expiry_hours = cache_expiry_hours
        self._schema = None
        self._controlled_vocabs = None
    
    def _is_cache_valid(self) -> bool:
        """Check if cached schema is still valid"""
        if not os.path.exists(self.cache_file):
            return False
        
        cache_time = datetime.fromtimestamp(os.path.getmtime(self.cache_file))
        expiry_time = cache_time + timedelta(hours=self.cache_expiry_hours)
        return datetime.now() < expiry_time
    
    def _fetch_schema(self) -> Dict:
        """Fetch schema from URL or cache"""
        if self._is_cache_valid():
            try:
                with open(self.cache_file, 'r') as f:
                    print(f"Loading cached schema from {self.cache_file}")
                    return json.load(f)
            except Exception as e:
                print(f"Error loading cache: {e}")
        
        print(f"Fetching schema from {self.jsonld_url}")
        try:
            response = requests.get(self.jsonld_url, timeout=30)
            response.raise_for_status()
            schema = response.json()
            
            # Cache the schema
            with open(self.cache_file, 'w') as f:
                json.dump(schema, f, indent=2)
            
            return schema
        except Exception as e:
            print(f"Error fetching schema: {e}")
            if os.path.exists(self.cache_file):
                print("Falling back to cached version")
                with open(self.cache_file, 'r') as f:
                    return json.load(f)
            raise
    
    def get_schema(self) -> Dict:
        """Get the complete schema"""
        if self._schema is None:
            self._schema = self._fetch_schema()
        return self._schema
    
    def extract_controlled_vocabularies(self) -> Dict[str, List[str]]:
        """Extract controlled vocabularies from the schema"""
        if self._controlled_vocabs is not None:
            return self._controlled_vocabs
        
        schema = self.get_schema()
        controlled_vocabs = {}
        
        # Group classes by their parent categories to form controlled vocabularies
        if '@graph' in schema:
            category_terms = {}
            
            for item in schema['@graph']:
                if item.get('@type') == 'rdfs:Class':
                    item_id = item.get('@id', '')
                    display_name = item.get('sms:displayName', '')
                    label = item.get('rdfs:label', '')
                    
                    # Extract the term name (use display name first, then label)
                    term_name = display_name or label
                    if not term_name:
                        continue
                    
                    # Determine category from parent classes
                    subclass_of = item.get('rdfs:subClassOf', [])
                    if isinstance(subclass_of, list):
                        for parent in subclass_of:
                            if isinstance(parent, dict) and '@id' in parent:
                                parent_id = parent['@id']
                                
                                # Extract category name from parent ID
                                if 'bts:' in parent_id:
                                    category = parent_id.replace('bts:', '').lower()
                                    
                                    # Map common categories to expected column names
                                    category_mapping = {
                                        'assay': 'assay',
                                        'specimentype': 'specimenType',
                                        'species': 'species',
                                        'sex': 'sex',
                                        'platform': 'platform',
                                        'datatype': 'dataType',
                                        'filetype': 'fileType',
                                        'diagnosis': 'diagnosis'
                                    }
                                    
                                    mapped_category = category_mapping.get(category, category)
                                    
                                    if mapped_category not in category_terms:
                                        category_terms[mapped_category] = []
                                    category_terms[mapped_category].append(term_name)
            
            # Also look for direct enum-like patterns
            for item in schema['@graph']:
                if item.get('@type') == 'rdfs:Class':
                    item_id = item.get('@id', '')
                    
                    # Look for Enum suffix patterns
                    if 'Enum' in item_id:
                        enum_base = item_id.replace('bts:', '').replace('Enum', '')
                        category = enum_base.lower()
                        
                        # Find all classes that are subclasses of this enum
                        enum_terms = []
                        for subitem in schema['@graph']:
                            if subitem.get('@type') == 'rdfs:Class':
                                subclass_of = subitem.get('rdfs:subClassOf', [])
                                if isinstance(subclass_of, list):
                                    for parent in subclass_of:
                                        if isinstance(parent, dict) and parent.get('@id') == item_id:
                                            term_name = subitem.get('sms:displayName') or subitem.get('rdfs:label', '')
                                            if term_name:
                                                enum_terms.append(term_name)
                        
                        if enum_terms:
                            category_mapping = {
                                'assay': 'assay',
                                'specimentype': 'specimenType',
                                'species': 'species',
                                'sex': 'sex',
                                'platform': 'platform'
                            }
                            mapped_category = category_mapping.get(category, category)
                            category_terms[mapped_category] = enum_terms
            
            controlled_vocabs = category_terms
        
        self._controlled_vocabs = controlled_vocabs
        return controlled_vocabs
    
    def get_vocabulary_for_column(self, column_name: str) -> List[str]:
        """Get controlled vocabulary terms for a specific column"""
        vocabs = self.extract_controlled_vocabularies()
        
        # Try exact match first
        if column_name in vocabs:
            return vocabs[column_name]
        
        # Try case-insensitive exact match
        for key, values in vocabs.items():
            if key.lower() == column_name.lower():
                return values
        
        # No match found - return empty list
        return []
    
    def get_column_description(self, column_name: str) -> str:
        """Get the description/definition for a specific column from the schema"""
        schema = self.get_schema()
        
        if '@graph' not in schema:
            return ""
        
        # Try to find a class definition that matches this column
        for item in schema['@graph']:
            if item.get('@type') == 'rdfs:Class':
                display_name = item.get('sms:displayName', '')
                label = item.get('rdfs:label', '')
                comment = item.get('rdfs:comment', '')
                
                # Match by display name or label
                if (display_name.lower() == column_name.lower() or 
                    label.lower() == column_name.lower()):
                    if comment and comment != 'TBD':
                        return comment
        
        return ""


class OpenRouterClient:
    """Client for OpenRouter API"""
    
    def __init__(self, api_key: str, model: str = "google/gemini-2-flash-exp"):
        self.api_key = api_key
        self.model = model
        self.base_url = "https://openrouter.ai/api/v1"
    
    def generate_mapping_suggestions(
        self, 
        row_contexts: List[str], 
        column_name: str,
        controlled_vocab: List[str],
        column_description: str = "",
        context: str = ""
    ) -> List[MappingSuggestion]:
        """Generate mapping suggestions using LLM"""
        
        if not controlled_vocab:
            return []
        
        prompt = self._build_mapping_prompt(row_contexts, column_name, controlled_vocab, column_description, context)
        
        try:
            response = requests.post(
                f"{self.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 4000
                },
                timeout=60
            )
            response.raise_for_status()
            
            result = response.json()
            content = result['choices'][0]['message']['content']
            
            return self._parse_mapping_response(content, row_contexts)
            
        except Exception as e:
            print(f"Error calling OpenRouter API: {e}")
            return []
    
    def _build_mapping_prompt(
        self, 
        row_contexts: List[str], 
        column_name: str, 
        controlled_vocab: List[str],
        column_description: str,
        context: str
    ) -> str:
        """Build the prompt for mapping suggestions"""
        
        vocab_list = "\n".join([f"- {term}" for term in controlled_vocab])
        context_list = "\n".join([f"Row {i+1}: {context}" for i, context in enumerate(row_contexts)])
        
        column_definition = f"\nCOLUMN DEFINITION: {column_description}" if column_description else ""
        
        prompt = f"""You are a metadata mapping expert. Your task is to determine the appropriate controlled vocabulary term(s) for the '{column_name}' column based on the data patterns.

TARGET COLUMN: {column_name}{column_definition}
{context}

CONTROLLED VOCABULARY TERMS (these are the ONLY valid options):
{vocab_list}

SAMPLE DATA ROWS (showing the complete context for each type of data):
{context_list}

Analyze the data and determine mapping strategy for the '{column_name}' column:

1. If this column should have the SAME value for all rows (e.g., assay, species, diagnosis for a homogeneous study), suggest ONE term.

2. If this column should vary by individual file characteristics (e.g., fileFormat based on file extensions, readPair based on specific read info), suggest different terms for different patterns and explain how to distinguish them.

3. Consider what the column definition tells you about whether values should be uniform or file-specific.

Respond in JSON format with the following structure:
[
  {{
    "input_value": "description of the data pattern or specific row context",
    "suggested_term": "matching controlled vocab term or NO_MATCH",
    "confidence": 0.95,
    "reasoning": "explanation based on the complete row context and data patterns"
  }}
]

IMPORTANT RULES:
1. Only use terms from the provided controlled vocabulary
2. Consider the ENTIRE row context, not just individual column values
3. Look for overall patterns in the data
4. Be conservative - if uncertain, use NO_MATCH
5. Provide confidence score from 0.0 to 1.0
6. Focus on what the data actually represents, not just text matching"""

        return prompt
    
    def _parse_mapping_response(self, content: str, row_contexts: List[str]) -> List[MappingSuggestion]:
        """Parse the LLM response into MappingSuggestion objects"""
        suggestions = []
        
        try:
            # Extract JSON from response
            start_idx = content.find('[')
            end_idx = content.rfind(']') + 1
            if start_idx == -1 or end_idx == 0:
                return suggestions
            
            json_content = content[start_idx:end_idx]
            mappings = json.loads(json_content)
            
            for mapping in mappings:
                if mapping.get('suggested_term') != 'NO_MATCH':
                    suggestions.append(MappingSuggestion(
                        original_value=mapping.get('input_value', ''),
                        suggested_term=mapping.get('suggested_term', ''),
                        confidence=float(mapping.get('confidence', 0.0)),
                        reasoning=mapping.get('reasoning', '')
                    ))
            
        except Exception as e:
            print(f"Error parsing LLM response: {e}")
            print(f"Response content: {content}")
        
        return suggestions


class MetadataMapper:
    """Main class for AI-powered metadata mapping"""
    
    def __init__(self, config_path: str = "config.yaml", creds_path: str = "creds.yaml"):
        self.config = self._load_config(config_path)
        self.creds = self._load_config(creds_path)
        
        # Initialize AI logger with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.ai_logger = AILogger(f"ai_api_calls_{timestamp}.log")
        
        # Track all cell mappings for Excel color coding
        self.cell_mappings: List[CellMapping] = []
        
        # Initialize components
        self.dictionary = NFMetadataDictionary(
            jsonld_url=self.config['metadata_dictionary']['jsonld_url'],
            cache_file=self.config['metadata_dictionary']['cache_file'],
            cache_expiry_hours=self.config['metadata_dictionary']['cache_expiry_hours']
        )
        
        self.llm_client = OpenRouterClient(
            api_key=self.creds['openrouter']['api_key'],
            model=self.config['llm']['model']
        )
    
    def _load_config(self, path: str) -> Dict:
        """Load YAML configuration file"""
        with open(path, 'r') as f:
            return yaml.safe_load(f)
    
    def process_csv_mapping(
        self, 
        input_csv_path: str, 
        target_csv_path: str, 
        output_csv_path: str
    ) -> None:
        """Main method to process CSV mapping with human-in-the-loop"""
        
        # Load CSVs
        input_df = pd.read_csv(input_csv_path)
        target_df = pd.read_csv(target_csv_path)
        
        print(f"Loaded input CSV with {len(input_df)} rows")
        print(f"Loaded target CSV template with {len(target_df)} rows")
        
        # Create comprehensive filename-to-row mapping upfront
        print("\n🔗 Creating filename-to-row mappings...")
        self.filename_to_input_mapping = self._create_comprehensive_filename_mapping(input_df, target_df)
        print(f"   Successfully mapped {len(self.filename_to_input_mapping)} target rows to input data")
        
        # Process each column in target CSV
        for column in target_df.columns:
            if self._should_skip_column(column, target_df):
                continue
            
            print(f"\n{'='*60}")
            print(f"Processing column: {column}")
            print(f"{'='*60}")
            
            # Get column description from schema first
            column_description = self.dictionary.get_column_description(column)
            
            # Get controlled vocabulary for this column
            controlled_vocab = self.dictionary.get_vocabulary_for_column(column)
            
            if not controlled_vocab:
                print(f"No controlled vocabulary found for column '{column}' - treating as freetext column...")
                self._process_freetext_column(input_df, target_df, column, column_description)
                continue
            
            print(f"Found {len(controlled_vocab)} controlled vocabulary terms")
            if column_description:
                print(f"Column definition: {column_description[:100]}...")
            
            # Get row contexts that could help map this column
            row_contexts = self._extract_relevant_input_values(input_df, column)
            
            if not row_contexts:
                print(f"No relevant input data found for column '{column}'")
                continue
            
            # Ask AI to determine if this column needs uniform or row-by-row mapping
            mapping_strategy = self._determine_mapping_strategy(row_contexts, column, controlled_vocab, column_description)
            
            if mapping_strategy['type'] == 'uniform':
                # Fill entire column with the same value
                uniform_value = mapping_strategy['value']
                confidence = mapping_strategy['confidence']
                reasoning = mapping_strategy['reasoning']
                
                print(f"Uniform mapping for '{column}': '{uniform_value}' (confidence: {confidence:.2f})")
                print(f"  Reasoning: {reasoning}")
                
                # Apply uniform value to all rows
                for target_idx in target_df.index:
                    target_df.at[target_idx, column] = uniform_value
                    self.cell_mappings.append(CellMapping(
                        row_idx=target_idx,
                        column_name=column,
                        mapped_value=uniform_value,
                        confidence=confidence,
                        reasoning=reasoning
                    ))
                
                print(f"Applied uniform value to {len(target_df)} rows")
                
            elif mapping_strategy['type'] == 'row_by_row':
                # Process in batches for row-by-row mapping
                print(f"Row-by-row mapping required for '{column}' - processing in batches...")
                self._apply_batch_row_mapping(input_df, target_df, column, controlled_vocab, column_description)
        
        # Save results as Excel with color coding
        excel_path = output_csv_path.replace('.csv', '.xlsx')
        self._save_excel_with_color_coding(target_df, excel_path)
        print(f"\nMapping complete! Results saved to: {excel_path}")
        
        # Also save as CSV for compatibility
        target_df.to_csv(output_csv_path, index=False)
        print(f"CSV version also saved to: {output_csv_path}")
        
        # Print AI usage summary
        self.ai_logger.print_summary()
    
    def _make_logged_api_call(self, function_name: str, prompt: str, max_tokens: int = 4000, 
                              response_format: dict = None, temperature: float = 0.1) -> dict:
        """Make an API call with logging and token tracking"""
        
        request_data = {
            "model": self.llm_client.model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": temperature,
            "max_tokens": max_tokens,
            "usage": {"include": True}  # Enable OpenRouter Usage Accounting
        }
        
        if response_format:
            request_data["response_format"] = response_format
        
        start_time = time.time()
        
        try:
            response = requests.post(
                f"{self.llm_client.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.llm_client.api_key}",
                    "Content-Type": "application/json"
                },
                json=request_data,
                timeout=60
            )
            response.raise_for_status()
            
            duration = time.time() - start_time
            result = response.json()
            
            # Log the API call
            self.ai_logger.log_api_call(
                function_name=function_name,
                model=self.llm_client.model,
                request_data=request_data,
                response_data=result,
                duration=duration
            )
            
            return result
            
        except Exception as e:
            duration = time.time() - start_time
            print(f"Error in {function_name}: {e}")
            
            # Log failed call
            error_response = {"error": str(e), "usage": {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}}
            self.ai_logger.log_api_call(
                function_name=f"{function_name}_FAILED",
                model=self.llm_client.model,
                request_data=request_data,
                response_data=error_response,
                duration=duration
            )
            raise
    
    def _should_skip_column(self, column: str, target_df: pd.DataFrame) -> bool:
        """Determine if a column should be skipped"""
        # Skip if column is already populated
        if not target_df[column].isna().all():
            populated_count = target_df[column].notna().sum()
            print(f"Column '{column}' already has {populated_count} populated values, skipping...")
            return True
        
        # Skip system columns
        system_columns = ['id', 'uuid', 'created_at', 'updated_at']
        if column.lower() in system_columns:
            print(f"Skipping system column: {column}")
            return True
        
        return False
    
    def _extract_relevant_input_values(self, input_df: pd.DataFrame, column_name: str) -> List[str]:
        """Extract input values that might be relevant for the target column"""
        row_contexts = []
        
        # Get a diverse sample of rows to capture variations in the data
        sample_size = min(15, len(input_df))  # Sample more rows to capture diversity
        
        # Try to get diverse samples by taking rows from different parts of the dataset
        if len(input_df) > sample_size:
            step = len(input_df) // sample_size
            sample_indices = list(range(0, len(input_df), step))[:sample_size]
            sample_rows = input_df.iloc[sample_indices]
        else:
            sample_rows = input_df.head(sample_size)
        
        # Look for key variation columns to ensure we capture diversity
        variation_columns = ['library_source', 'experiment_title', 'sample_title', 'library_strategy']
        
        # If we have variation in key columns, make sure to sample from different groups
        for var_col in variation_columns:
            if var_col in input_df.columns:
                unique_values = input_df[var_col].dropna().unique()
                if len(unique_values) > 1:
                    # Sample at least one row from each unique value
                    diverse_samples = []
                    for unique_val in unique_values[:10]:  # Limit to 10 different types
                        matching_rows = input_df[input_df[var_col] == unique_val]
                        if len(matching_rows) > 0:
                            diverse_samples.append(matching_rows.iloc[0])
                    
                    if diverse_samples:
                        sample_rows = pd.DataFrame(diverse_samples)
                        break
        
        for idx, row in sample_rows.iterrows():
            # Create a context string that describes this row
            row_context_parts = []
            for col, value in row.items():
                if pd.notna(value):
                    # For filename-related columns, always include filenames
                    if column_name.lower() in ['fileformat', 'filename', 'file_name'] and 'file' in col.lower():
                        row_context_parts.append(f"{col}: {value}")
                    elif not self._looks_like_filename_or_id(str(value)):
                        row_context_parts.append(f"{col}: {value}")
            
            if row_context_parts:
                row_context = "; ".join(row_context_parts)
                row_contexts.append(row_context)
        
        return row_contexts
    
    def _looks_like_filename_or_id(self, value: str) -> bool:
        """Check if a value looks like a filename or ID that shouldn't be mapped"""
        value = value.lower()
        
        # File extensions
        if any(value.endswith(ext) for ext in ['.fastq', '.bam', '.vcf', '.csv', '.txt', '.gz']):
            return True
        
        # UUIDs or long alphanumeric strings
        if len(value) > 20 and (value.replace('-', '').replace('_', '').isalnum()):
            return True
        
        # Pure numbers
        if value.replace('.', '').replace('-', '').isdigit():
            return True
        
        return False
    
    def _validate_suggestions(
        self, 
        suggestions: List[MappingSuggestion], 
        controlled_vocab: List[str]
    ) -> List[MappingSuggestion]:
        """Validate suggestions against controlled vocabulary"""
        for suggestion in suggestions:
            suggestion.is_valid = suggestion.suggested_term in controlled_vocab
        
        return [s for s in suggestions if s.is_valid]
    
    
    def _apply_mappings_to_target(
        self, 
        input_df: pd.DataFrame, 
        target_df: pd.DataFrame, 
        column_name: str,
        column_mapping: ColumnMapping, 
        row_contexts: List[str]
    ) -> None:
        """Apply the approved mappings to the target CSV"""
        
        if not column_mapping.selected_mapping:
            print(f"No mappings to apply for column: {column_name}")
            return
        
        mappings_applied = 0
        
        # Process each row individually to find the best mapping based on its specific data
        default_mapping = None
        default_confidence = 0.0
        default_reasoning = ""
        
        # Get a default mapping if all suggestions point to the same term
        if len(column_mapping.selected_mapping) == 1:
            default_mapping = list(column_mapping.selected_mapping.values())[0]
            for suggestion in column_mapping.suggestions:
                if suggestion.suggested_term == default_mapping:
                    default_confidence = suggestion.confidence
                    default_reasoning = suggestion.reasoning
                    break
        
        for target_idx, target_row in target_df.iterrows():
            # Try to match by filename first
            target_filename = target_row.get('fileName', target_row.get('filename', target_row.get('file_name', target_row.get('Filename', ''))))
            
            if not target_filename:
                continue
                
            # Find matching row in input CSV by filename
            input_row = self._find_matching_input_row(input_df, target_filename)
            
            if input_row is None:
                # If no matching input row, use default mapping if available
                if default_mapping:
                    target_df.at[target_idx, column_name] = default_mapping
                    self.cell_mappings.append(CellMapping(
                        row_idx=target_idx,
                        column_name=column_name,
                        mapped_value=default_mapping,
                        confidence=default_confidence,
                        reasoning=default_reasoning
                    ))
                    mappings_applied += 1
                continue
            
            # Find the best mapping for this specific row's data
            best_mapping, best_confidence, best_reasoning = self._find_best_mapping_for_row(input_row, column_mapping)
            
            # Apply the best mapping found for this row, or use default if no specific match
            final_mapping = best_mapping if best_mapping else default_mapping
            final_confidence = best_confidence if best_mapping else default_confidence
            final_reasoning = best_reasoning if best_mapping else default_reasoning
            
            if final_mapping:
                target_df.at[target_idx, column_name] = final_mapping
                # Track this cell mapping for color coding
                self.cell_mappings.append(CellMapping(
                    row_idx=target_idx,
                    column_name=column_name,
                    mapped_value=final_mapping,
                    confidence=final_confidence,
                    reasoning=final_reasoning
                ))
                mappings_applied += 1
        
        print(f"Applied {mappings_applied} mappings to column '{column_name}'")
    
    def _has_corresponding_input_file(self, input_df: pd.DataFrame, target_filename: str) -> bool:
        """Check if a target filename has a corresponding row in the input data"""
        for fname_col in ['file_name', 'fileName', 'filename', 'Filename']:
            if fname_col in input_df.columns:
                if target_filename in input_df[fname_col].astype(str).values:
                    return True
        return False
    
    def _find_matching_input_row(self, input_df: pd.DataFrame, target_filename: str):
        """Find the input row that matches the target filename using AI-powered matching"""
        
        # First try the existing rule-based approach for speed
        quick_match = self._try_rule_based_matching(input_df, target_filename)
        if quick_match is not None:
            return quick_match
            
        # If rule-based fails, use AI-powered matching
        return self._ai_powered_filename_matching(input_df, target_filename)
    
    def _try_rule_based_matching(self, input_df: pd.DataFrame, target_filename: str):
        """Try fast rule-based matching first"""
        # First try exact filename matching
        for fname_col in ['file_name', 'fileName', 'filename', 'Filename']:
            if fname_col in input_df.columns:
                exact_matches = input_df[input_df[fname_col].astype(str) == target_filename]
                if not exact_matches.empty:
                    return exact_matches.iloc[0]
        
        # Extract potential identifiers from target filename
        base_filename = target_filename.replace('.fastq.gz', '').replace('.fastq', '').replace('.bam', '').replace('.vcf', '')
        
        # Try matching on run_accession, experiment_accession patterns in filename
        filename_parts = base_filename.split('_')
        for part in filename_parts:
            if len(part) > 6:  # Skip short parts
                # Check run_accession
                if 'run_accession' in input_df.columns:
                    accession_matches = input_df[input_df['run_accession'].astype(str) == part]
                    if not accession_matches.empty:
                        return accession_matches.iloc[0]
                
                # Check experiment_accession  
                if 'experiment_accession' in input_df.columns:
                    exp_matches = input_df[input_df['experiment_accession'].astype(str) == part]
                    if not exp_matches.empty:
                        return exp_matches.iloc[0]
        
        return None
    
    def _ai_powered_filename_matching(self, input_df: pd.DataFrame, target_filename: str):
        """Use AI to intelligently match target filename to input data row"""
        
        # Create a summary of available input rows (limit to first 10 for efficiency)
        sample_rows = input_df.head(10)
        
        # Create row summaries for AI
        row_summaries = []
        for idx, row in sample_rows.iterrows():
            # Include key identifying fields
            summary_fields = []
            for col in ['run_accession', 'experiment_accession', 'sample_accession', 'biosample', 
                       'experiment_title', 'sample_title', 'file_name']:
                if col in row and pd.notna(row[col]) and str(row[col]).strip():
                    summary_fields.append(f"{col}: {row[col]}")
            
            row_summaries.append({
                'row_index': idx,
                'summary': "; ".join(summary_fields[:6])  # Limit to avoid token overflow
            })
        
        if not row_summaries:
            return None
            
        prompt = f"""You are an expert at matching filenames to metadata rows. 

TARGET FILENAME: {target_filename}

AVAILABLE INPUT ROWS:
{chr(10).join(f"Row {r['row_index']}: {r['summary']}" for r in row_summaries)}

TASK: Determine which input row (by row_index) best matches the target filename.

Look for patterns like:
- Accession numbers (SRR, SRX, SAMN) in the filename matching row data
- Sample identifiers or experimental conditions in filename matching experiment_title
- File naming patterns that correspond to the metadata

If you find a clear match, respond with ONLY the row_index number.
If no clear match exists, respond with "NO_MATCH".

Response:"""

        try:
            # Use AI to determine the best matching row
            response = self._make_logged_api_call(
                function_name="_ai_powered_filename_matching",
                prompt=prompt,
                max_tokens=50,
                temperature=0.1
            )
            
            ai_response = response['choices'][0]['message']['content'].strip()
            
            # Parse AI response
            if ai_response == "NO_MATCH":
                return None
            
            # Try to extract row index
            try:
                row_idx = int(ai_response)
                if row_idx in input_df.index:
                    return input_df.loc[row_idx]
            except (ValueError, KeyError):
                pass
                
        except Exception as e:
            print(f"    AI filename matching failed: {e}")
        
        return None
    
    def _create_comprehensive_filename_mapping(self, input_df: pd.DataFrame, target_df: pd.DataFrame):
        """Create a comprehensive mapping between target filenames and input data rows upfront"""
        
        filename_mapping = {}
        
        # Check if AI-powered filename matching is enabled
        ai_filename_matching = self.config.get('llm', {}).get('ai_powered_filename_matching', True)
        
        if ai_filename_matching:
            # Use AI to create intelligent mappings for all target rows at once
            filename_mapping = self._create_ai_comprehensive_mapping(input_df, target_df)
        
        # Fill in any missing mappings with rule-based approach
        for target_idx, target_row in target_df.iterrows():
            if target_idx not in filename_mapping:
                filename = str(target_row.get('Filename', '')).strip()
                input_row = self._try_rule_based_matching(input_df, filename)
                if input_row is not None:
                    filename_mapping[target_idx] = input_row
        
        return filename_mapping
    
    def _create_ai_comprehensive_mapping(self, input_df: pd.DataFrame, target_df: pd.DataFrame):
        """Use AI to create comprehensive filename-to-row mappings for all target rows"""
        
        # Extract all target filenames
        target_info = []
        for idx, row in target_df.iterrows():
            filename = str(row.get('Filename', '')).strip()
            if filename:  # Only include rows with filenames
                target_info.append({
                    'target_idx': idx,
                    'filename': filename
                })
        
        if not target_info:
            return {}
        
        # Create input data summary (limit for efficiency but include more than before)
        input_summary = []
        for idx, row in input_df.head(50).iterrows():  # Increased from 20 to 50
            summary_fields = []
            for col in ['run_accession', 'experiment_accession', 'sample_accession', 'biosample',
                       'experiment_title', 'sample_title', 'file_name']:
                if col in row and pd.notna(row[col]) and str(row[col]).strip():
                    summary_fields.append(f"{col}: {row[col]}")
            
            if summary_fields:
                input_summary.append({
                    'input_idx': idx,
                    'summary': "; ".join(summary_fields[:5])
                })
        
        if not input_summary:
            return {}
        
        # Process in chunks to avoid token limits
        chunk_size = 20  # Process 20 target files at a time
        filename_mapping = {}
        
        for i in range(0, len(target_info), chunk_size):
            chunk_targets = target_info[i:i + chunk_size]
            chunk_mapping = self._process_ai_mapping_chunk(chunk_targets, input_summary)
            filename_mapping.update(chunk_mapping)
        
        return filename_mapping
    
    def _process_ai_mapping_chunk(self, target_chunk: list, input_summary: list):
        """Process a chunk of target filenames with AI mapping"""
        
        target_files_str = "\n".join(f"Target {t['target_idx']}: {t['filename']}" for t in target_chunk)
        input_data_str = "\n".join(f"Input {i['input_idx']}: {i['summary']}" for i in input_summary)
        
        prompt = f"""You are an expert at mapping target filenames to input metadata rows.

TARGET FILENAMES:
{target_files_str}

AVAILABLE INPUT DATA:
{input_data_str}

TASK: Map each target filename to the best matching input row index.

Look for patterns like:
- Accession numbers (SRR, SRX, SAMN) in filenames matching input data
- Sample identifiers or experimental conditions in filenames matching experiment titles
- File naming patterns that correspond to the metadata

Respond with mappings in this exact format:
target_idx:input_idx
target_idx:input_idx
...

If no match exists for a target, use: target_idx:NO_MATCH

Response:"""

        try:
            response = self._make_logged_api_call(
                function_name="_process_ai_mapping_chunk",
                prompt=prompt,
                max_tokens=500,
                temperature=0.1
            )
            
            ai_response = response['choices'][0]['message']['content'].strip()
            return self._parse_comprehensive_ai_mappings(ai_response, target_chunk, input_summary)
            
        except Exception as e:
            print(f"    AI comprehensive mapping failed for chunk: {e}")
            return {}
    
    def _parse_comprehensive_ai_mappings(self, ai_response: str, target_chunk: list, input_summary: list):
        """Parse AI response and create filename mappings"""
        filename_mapping = {}
        
        # Create input index to row mapping for quick lookup
        input_idx_to_row = {}
        for input_info in input_summary:
            input_idx = input_info['input_idx']
            # We need to get the actual row from the DataFrame - we'll do this in the batch processing
            input_idx_to_row[input_idx] = input_idx  # Store the index for now
        
        for line in ai_response.strip().split('\n'):
            if ':' in line:
                try:
                    target_part, input_part = line.split(':', 1)
                    target_idx = int(target_part.strip())
                    input_part = input_part.strip()
                    
                    if input_part != "NO_MATCH":
                        input_idx = int(input_part)
                        if input_idx in input_idx_to_row:
                            # Store the input index - we'll resolve to actual row later
                            filename_mapping[target_idx] = input_idx
                            
                except (ValueError, IndexError):
                    continue
        
        return filename_mapping
    
    def _create_batch_mappings_from_precomputed(self, input_df: pd.DataFrame, batch_target_rows: pd.DataFrame):
        """Create batch mappings using pre-computed filename-to-row mappings"""
        
        batch_mappings = []
        
        for target_idx, target_row in batch_target_rows.iterrows():
            filename = str(target_row.get('Filename', '')).strip()
            
            # Check if we have a pre-computed mapping for this target row
            if hasattr(self, 'filename_to_input_mapping') and target_idx in self.filename_to_input_mapping:
                input_idx_or_row = self.filename_to_input_mapping[target_idx]
                
                # Handle both cases: input index (from AI mapping) or actual row (from rule-based)
                if isinstance(input_idx_or_row, int):
                    # It's an index from AI mapping
                    if input_idx_or_row in input_df.index:
                        input_row = input_df.loc[input_idx_or_row]
                    else:
                        continue  # Skip if index is invalid
                else:
                    # It's an actual row from rule-based matching
                    input_row = input_idx_or_row
                
                row_context = self._create_row_context_string(input_row)
                batch_mappings.append({
                    'target_idx': target_idx,
                    'filename': filename,
                    'context': row_context
                })
        
        return batch_mappings
    
    def _create_ai_powered_batch_mappings(self, input_df: pd.DataFrame, batch_target_rows: pd.DataFrame, 
                                         column_name: str, column_description: str):
        """Use AI to create intelligent mappings between target filenames and input data"""
        
        # Check if AI-powered filename matching is enabled
        ai_filename_matching = self.config.get('llm', {}).get('ai_powered_filename_matching', True)
        if not ai_filename_matching:
            # Use rule-based matching only, no fallback to representative data
            batch_mappings = []
            for target_idx, target_row in batch_target_rows.iterrows():
                filename = str(target_row.get('Filename', '')).strip()
                input_row = self._try_rule_based_matching(input_df, filename)
                
                if input_row is not None:
                    row_context = self._create_row_context_string(input_row)
                    batch_mappings.append({
                        'target_idx': target_idx,
                        'filename': filename,
                        'context': row_context
                    })
                # If no match found, skip (leave blank)
            return batch_mappings
        
        # Extract target filenames
        target_filenames = []
        for idx, row in batch_target_rows.iterrows():
            filename = str(row.get('Filename', '')).strip()
            target_filenames.append({
                'target_idx': idx,
                'filename': filename
            })
        
        if not target_filenames:
            return []
            
        # Create input data summary (limit for efficiency)
        input_summary = []
        for idx, row in input_df.head(20).iterrows():
            summary_fields = []
            for col in ['run_accession', 'experiment_accession', 'sample_accession', 'biosample',
                       'experiment_title', 'sample_title', 'file_name']:
                if col in row and pd.notna(row[col]) and str(row[col]).strip():
                    summary_fields.append(f"{col}: {row[col]}")
            
            if summary_fields:
                input_summary.append({
                    'input_idx': idx,
                    'summary': "; ".join(summary_fields[:5])
                })
        
        if not input_summary:
            # No input data available - return empty mappings (blank values)
            return []
            
        # Create AI prompt for batch mapping
        target_files_str = "\n".join(f"Target {t['target_idx']}: {t['filename']}" for t in target_filenames[:10])
        input_data_str = "\n".join(f"Input {i['input_idx']}: {i['summary']}" for i in input_summary[:15])
        
        prompt = f"""You are an expert at mapping target filenames to input metadata for the column '{column_name}'.

COLUMN: {column_name}
DEFINITION: {column_description}

TARGET FILENAMES:
{target_files_str}

AVAILABLE INPUT DATA:
{input_data_str}

TASK: Map each target filename to the best matching input row index.

Look for patterns like:
- Accession numbers (SRR, SRX, SAMN) in filenames matching input data
- Sample identifiers or experimental conditions in filenames matching experiment titles
- File naming patterns that correspond to the metadata

Respond with mappings in this format:
target_idx:input_idx
target_idx:input_idx
...

If no match exists for a target, use: target_idx:NO_MATCH

Response:"""

        try:
            response = self._make_logged_api_call(
                function_name="_create_ai_powered_batch_mappings",
                prompt=prompt,
                max_tokens=200,
                temperature=0.1
            )
            
            ai_response = response['choices'][0]['message']['content'].strip()
            return self._parse_ai_batch_mappings(ai_response, input_df, target_filenames)
            
        except Exception as e:
            print(f"    AI batch mapping failed: {e}")
            # Return empty mappings - leave blank rather than use fallback
            return []
    
    def _parse_ai_batch_mappings(self, ai_response: str, input_df: pd.DataFrame, target_filenames: list):
        """Parse AI response and create batch mappings"""
        batch_mappings = []
        
        for line in ai_response.strip().split('\n'):
            if ':' in line:
                try:
                    target_part, input_part = line.split(':', 1)
                    target_idx = int(target_part.strip())
                    input_part = input_part.strip()
                    
                    if input_part == "NO_MATCH":
                        # Skip this target - leave blank rather than use fallback data
                        continue
                    else:
                        input_idx = int(input_part)
                        if input_idx in input_df.index:
                            input_row = input_df.loc[input_idx]
                            row_context = self._create_row_context_string(input_row)
                            filename = next((t['filename'] for t in target_filenames if t['target_idx'] == target_idx), '')
                            batch_mappings.append({
                                'target_idx': target_idx,
                                'filename': filename,
                                'context': row_context
                            })
                except (ValueError, IndexError):
                    continue
        
        # If AI parsing failed, leave blank rather than use fallback data
        # Empty batch_mappings will result in blank values for unmatched targets
                
        return batch_mappings
    
    
    def _find_best_mapping_for_row(self, input_row, column_mapping: ColumnMapping):
        """Ask the AI directly what value should be used for this specific row and column"""
        
        # Special handling for file format - check file extensions (this is deterministic)
        if column_mapping.column_name.lower() == 'fileformat':
            return self._map_file_format_from_row(input_row, column_mapping)
        
        # Create a clean row context
        row_context = self._create_row_context_string(input_row)
        
        # Get controlled vocabulary for this column
        controlled_vocab = [sugg.suggested_term for sugg in column_mapping.suggestions]
        
        # Get column description
        column_description = self.dictionary.get_column_description(column_mapping.column_name)
        
        # Ask the AI directly what value to use for this specific row
        prompt = f"""You are filling in a metadata CSV. For a specific row of data, determine the correct value for the '{column_mapping.column_name}' column.

TARGET COLUMN: {column_mapping.column_name}
COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

AVAILABLE VALUES (controlled vocabulary):
{', '.join(controlled_vocab) if controlled_vocab else 'No controlled vocabulary available'}

INPUT ROW DATA: {row_context}

Based on this specific row's data, what should the value be for the '{column_mapping.column_name}' column? 

Respond in this JSON format:
{{
  "value": "exact_value_from_controlled_vocabulary_or_blank",
  "confidence": 0.95,
  "reasoning": "brief explanation"
}}

If no appropriate value can be determined from the available options, use "value": "" (blank)."""

        try:
            # Make direct API call using the same pattern as the LLM client
            response = requests.post(
                f"{self.llm_client.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.llm_client.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.llm_client.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 200
                },
                timeout=30
            )
            response.raise_for_status()
            
            result = response.json()
            response_content = result['choices'][0]['message']['content']
            
            # Parse the response
            import json
            try:
                result = json.loads(response_content.strip())
                value = result.get('value', '')
                confidence = float(result.get('confidence', 0.0))
                reasoning = result.get('reasoning', 'AI direct mapping')
                
                return value, confidence, reasoning
                
            except json.JSONDecodeError:
                # Fallback: try to extract value from text response
                if controlled_vocab:
                    for vocab_term in controlled_vocab:
                        if vocab_term.lower() in response_content.lower():
                            return vocab_term, 0.7, "Fallback text extraction"
                
                return None, 0.0, "Could not parse AI response"
                
        except Exception as e:
            print(f"Error in direct AI mapping: {e}")
            return None, 0.0, f"Error: {str(e)}"
    
    def _row_matches_pattern(self, input_row, pattern_description: str, column_name: str, llm_client) -> bool:
        """Ask the AI if this row matches the pattern - completely dynamic"""
        # For now, let's use a simpler approach to avoid excessive API calls
        # TODO: Implement efficient AI-driven pattern matching
        return self._simple_keyword_fallback(input_row, pattern_description)
    
    def _create_row_context_string(self, input_row) -> str:
        """Create a clean context string for a single row"""
        context_parts = []
        for col, value in input_row.items():
            if pd.notna(value) and not self._looks_like_filename_or_id(str(value)):
                context_parts.append(f"{col}: {value}")
        return "; ".join(context_parts)
    
    def _simple_keyword_fallback(self, input_row, pattern_description: str) -> bool:
        """Simple fallback when AI call fails"""
        pattern_keywords = set(pattern_description.lower().split())
        
        for col, value in input_row.items():
            if pd.notna(value):
                value_keywords = set(str(value).lower().split())
                if len(pattern_keywords.intersection(value_keywords)) >= 2:
                    return True
        return False
    
    def _determine_mapping_strategy(self, row_contexts: List[str], column_name: str, controlled_vocab: List[str], column_description: str) -> dict:
        """Ask AI whether this column should be filled uniformly or row-by-row"""
        
        # Sample up to 50 rows for analysis
        sample_contexts = row_contexts[:50]
        
        prompt = f"""You are analyzing how to fill a metadata column. Look at sample data and determine the strategy.

TARGET COLUMN: {column_name}
COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

AVAILABLE VALUES (controlled vocabulary):
{', '.join(controlled_vocab) if controlled_vocab else 'No controlled vocabulary available'}

SAMPLE INPUT DATA ({len(sample_contexts)} rows):
{chr(10).join(f"Row {i+1}: {context}" for i, context in enumerate(sample_contexts))}

Based on this data, should the '{column_name}' column be:
1. UNIFORM - All rows get the same value (e.g., all samples are from "Homo sapiens")
2. ROW_BY_ROW - Different rows need different values (e.g., some are "RNA-seq", others are "single-cell RNA-seq")

If UNIFORM, what single value should be used for all rows?
If ROW_BY_ROW, explain why rows differ.

Respond in this JSON format:
{{
  "type": "uniform" or "row_by_row",
  "value": "single_value_if_uniform_or_empty_if_row_by_row",
  "confidence": 0.95,
  "reasoning": "explanation of decision"
}}"""

        response_format = {
            "type": "json_schema",
            "json_schema": {
                "name": "mapping_strategy",
                "strict": True,
                "schema": {
                    "type": "object",
                    "properties": {
                        "type": {
                            "type": "string",
                            "enum": ["uniform", "row_by_row"],
                            "description": "Whether all rows should get the same value or need individual analysis"
                        },
                        "value": {
                            "type": "string",
                            "description": "The single value if uniform, empty string if row_by_row"
                        },
                        "confidence": {
                            "type": "number",
                            "minimum": 0,
                            "maximum": 1,
                            "description": "Confidence score between 0 and 1"
                        },
                        "reasoning": {
                            "type": "string",
                            "description": "Brief explanation of the decision"
                        }
                    },
                    "required": ["type", "value", "confidence", "reasoning"],
                    "additionalProperties": False
                }
            }
        }
        
        try:
            result = self._make_logged_api_call(
                function_name="determine_mapping_strategy",
                prompt=prompt,
                max_tokens=500,
                response_format=response_format
            )
            
            response_content = result['choices'][0]['message']['content'].strip()
            
            # With structured outputs, this should always be valid JSON
            strategy = json.loads(response_content)
            return {
                'type': strategy['type'],
                'value': strategy['value'],
                'confidence': float(strategy['confidence']),
                'reasoning': strategy['reasoning']
            }
                
        except Exception as e:
            print(f"Error determining mapping strategy: {e}")
            # If structured outputs failed, try without them
            if "structured" in str(e).lower() or "json_schema" in str(e).lower():
                print("  Structured outputs not supported, falling back to regular prompting...")
                return self._determine_mapping_strategy_fallback(row_contexts, column_name, controlled_vocab, column_description)
            return {
                'type': 'row_by_row',
                'value': '',
                'confidence': 0.5,
                'reasoning': f'Error: {str(e)}'
            }
    
    def _determine_mapping_strategy_fallback(self, row_contexts: List[str], column_name: str, controlled_vocab: List[str], column_description: str) -> dict:
        """Fallback method without structured outputs for models that don't support them"""
        
        sample_contexts = row_contexts[:50]
        
        prompt = f"""Analyze how to fill a metadata column. Respond with EXACTLY this format:

TARGET COLUMN: {column_name}
COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

AVAILABLE VALUES: {', '.join(controlled_vocab) if controlled_vocab else 'No controlled vocabulary available'}

SAMPLE DATA ({len(sample_contexts)} rows):
{chr(10).join(f"Row {i+1}: {context}" for i, context in enumerate(sample_contexts))}

DECISION: [UNIFORM or ROW_BY_ROW]
VALUE: [single_value_if_uniform_or_VARIES_if_row_by_row]
CONFIDENCE: [0.0-1.0]
REASONING: [explanation]

Fill in the brackets with your analysis."""

        try:
            response = requests.post(
                f"{self.llm_client.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.llm_client.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.llm_client.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 500
                },
                timeout=60
            )
            response.raise_for_status()
            
            result = response.json()
            response_content = result['choices'][0]['message']['content']
            
            # Parse the structured text response
            return self._parse_strategy_response(response_content)
                
        except Exception as e:
            print(f"Error in fallback strategy: {e}")
            return {
                'type': 'row_by_row',
                'value': '',
                'confidence': 0.5,
                'reasoning': f'Fallback error: {str(e)}'
            }
    
    def _parse_strategy_response(self, response_content: str) -> dict:
        """Parse the structured text response format"""
        lines = response_content.strip().split('\n')
        
        strategy_type = 'row_by_row'
        value = ''
        confidence = 0.5
        reasoning = 'Could not parse response'
        
        for line in lines:
            line = line.strip()
            if line.startswith('DECISION:'):
                decision = line.replace('DECISION:', '').strip().strip('[]')
                strategy_type = 'uniform' if 'UNIFORM' in decision.upper() else 'row_by_row'
            elif line.startswith('VALUE:'):
                value = line.replace('VALUE:', '').strip().strip('[]')
                if value.upper() == 'VARIES':
                    value = ''
            elif line.startswith('CONFIDENCE:'):
                try:
                    confidence = float(line.replace('CONFIDENCE:', '').strip().strip('[]'))
                except:
                    confidence = 0.5
            elif line.startswith('REASONING:'):
                reasoning = line.replace('REASONING:', '').strip().strip('[]')
        
        return {
            'type': strategy_type,
            'value': value,
            'confidence': confidence,
            'reasoning': reasoning
        }
    
    def _apply_batch_row_mapping(self, input_df: pd.DataFrame, target_df: pd.DataFrame, column_name: str, controlled_vocab: List[str], column_description: str):
        """Apply row-by-row mapping in batches of 50"""
        
        batch_size = 50
        total_batches = (len(target_df) + batch_size - 1) // batch_size
        
        for batch_num in range(total_batches):
            start_idx = batch_num * batch_size
            end_idx = min(start_idx + batch_size, len(target_df))
            batch_target_rows = target_df.iloc[start_idx:end_idx]
            
            print(f"  Processing batch {batch_num + 1}/{total_batches} (rows {start_idx + 1}-{end_idx})")
            
            # Use pre-computed filename mappings for this batch
            batch_mappings = self._create_batch_mappings_from_precomputed(
                input_df, batch_target_rows)
            
            if not batch_mappings:
                continue
                
            # Ask AI to map this batch
            batch_results = self._get_batch_mappings(batch_mappings, column_name, controlled_vocab, column_description)
            
            # Apply the results
            for result in batch_results:
                target_idx = result['target_idx']
                value = result['value']
                confidence = result['confidence']
                reasoning = result['reasoning']
                
                if value:
                    target_df.at[target_idx, column_name] = value
                    self.cell_mappings.append(CellMapping(
                        row_idx=target_idx,
                        column_name=column_name,
                        mapped_value=value,
                        confidence=confidence,
                        reasoning=reasoning
                    ))
        
        mapped_count = sum(1 for mapping in self.cell_mappings if mapping.column_name == column_name)
        print(f"Applied {mapped_count} row-specific mappings to column '{column_name}'")
    
    def _get_batch_mappings(self, batch_mappings: List[dict], column_name: str, controlled_vocab: List[str], column_description: str) -> List[dict]:
        """Ask AI to determine values for a batch of rows"""
        
        rows_text = "\n".join([
            f"Row {i+1} (file: {mapping['filename']}): {mapping['context']}"
            for i, mapping in enumerate(batch_mappings)
        ])
        
        prompt = f"""You are filling in a metadata column for multiple rows. For each row, determine the correct value.

TARGET COLUMN: {column_name}
COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

AVAILABLE VALUES (controlled vocabulary):
{', '.join(controlled_vocab) if controlled_vocab else 'No controlled vocabulary available'}

ROWS TO PROCESS:
{rows_text}

For each row, determine what value should go in the '{column_name}' column.

Respond in this JSON format:
{{
  "mappings": [
    {{"row": 1, "value": "exact_value_from_vocab_or_blank", "confidence": 0.95, "reasoning": "brief explanation"}},
    {{"row": 2, "value": "exact_value_from_vocab_or_blank", "confidence": 0.90, "reasoning": "brief explanation"}},
    ...
  ]
}}

Use "" (blank) for value if no appropriate mapping can be determined."""

        try:
            response = requests.post(
                f"{self.llm_client.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.llm_client.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.llm_client.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 4000,  # Increase token limit
                    "response_format": {
                        "type": "json_schema",
                        "json_schema": {
                            "name": "batch_mappings",
                            "strict": True,
                            "schema": {
                                "type": "object",
                                "properties": {
                                    "mappings": {
                                        "type": "array",
                                        "items": {
                                            "type": "object",
                                            "properties": {
                                                "row": {
                                                    "type": "integer",
                                                    "minimum": 1,
                                                    "description": "Row number being mapped"
                                                },
                                                "value": {
                                                    "type": "string",
                                                    "description": "The mapped value from controlled vocabulary, or empty string if no mapping"
                                                },
                                                "confidence": {
                                                    "type": "number",
                                                    "minimum": 0,
                                                    "maximum": 1,
                                                    "description": "Confidence score between 0 and 1"
                                                },
                                                "reasoning": {
                                                    "type": "string",
                                                    "description": "Brief explanation for this mapping"
                                                }
                                            },
                                            "required": ["row", "value", "confidence", "reasoning"],
                                            "additionalProperties": False
                                        }
                                    }
                                },
                                "required": ["mappings"],
                                "additionalProperties": False
                            }
                        }
                    }
                },
                timeout=60
            )
            response.raise_for_status()
            
            result = response.json()
            response_content = result['choices'][0]['message']['content']
            
            # Debug: Print the response to see what we're getting
            print(f"    Raw response length: {len(response_content)}")
            print(f"    First 200 chars: {response_content[:200]}")
            print(f"    Last 200 chars: {response_content[-200:]}")
            
            # With structured outputs, this should always be valid JSON
            batch_result = json.loads(response_content.strip())
            mappings = batch_result['mappings']
            
            results = []
            for mapping in mappings:
                row_num = mapping['row'] - 1  # Convert to 0-based index
                if 0 <= row_num < len(batch_mappings):
                    results.append({
                        'target_idx': batch_mappings[row_num]['target_idx'],
                        'value': mapping['value'],
                        'confidence': float(mapping['confidence']),
                        'reasoning': mapping['reasoning']
                    })
            
            return results
                
        except Exception as e:
            print(f"Error in batch mapping: {e}")
            # If structured outputs failed, try without them
            if "structured" in str(e).lower() or "json_schema" in str(e).lower():
                print("    Falling back to regular prompting for batch mapping...")
                return self._get_batch_mappings_fallback(batch_mappings, column_name, controlled_vocab, column_description)
            return []
    
    def _get_batch_mappings_fallback(self, batch_mappings: List[dict], column_name: str, controlled_vocab: List[str], column_description: str) -> List[dict]:
        """Fallback batch mapping without structured outputs"""
        
        rows_text = "\n".join([
            f"Row {i+1} (file: {mapping['filename']}): {mapping['context']}"
            for i, mapping in enumerate(batch_mappings)
        ])
        
        prompt = f"""Map values for multiple rows. Respond in EXACTLY this format:

TARGET COLUMN: {column_name}
DEFINITION: {column_description if column_description else 'No definition available'}
VOCABULARY: {', '.join(controlled_vocab) if controlled_vocab else 'No controlled vocabulary available'}

ROWS:
{rows_text}

MAPPINGS:
Row 1: [value_or_blank]
Row 2: [value_or_blank]
...

Use exact values from vocabulary or leave blank."""

        try:
            response = requests.post(
                f"{self.llm_client.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.llm_client.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": self.llm_client.model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "max_tokens": 1500
                },
                timeout=60
            )
            response.raise_for_status()
            
            result = response.json()
            response_content = result['choices'][0]['message']['content']
            
            # Parse the structured text response
            return self._parse_batch_response(response_content, batch_mappings)
                
        except Exception as e:
            print(f"Error in fallback batch mapping: {e}")
            return []
    
    def _parse_batch_response(self, response_content: str, batch_mappings: List[dict]) -> List[dict]:
        """Parse the structured text batch response"""
        lines = response_content.strip().split('\n')
        results = []
        
        for line in lines:
            line = line.strip()
            if line.startswith('Row ') and ':' in line:
                try:
                    # Extract row number and value
                    parts = line.split(':', 1)
                    row_part = parts[0].strip()
                    value_part = parts[1].strip().strip('[]')
                    
                    # Get row number
                    row_num = int(row_part.replace('Row', '').strip()) - 1  # Convert to 0-based
                    
                    if 0 <= row_num < len(batch_mappings):
                        results.append({
                            'target_idx': batch_mappings[row_num]['target_idx'],
                            'value': value_part if value_part.lower() not in ['blank', 'empty', 'none', ''] else '',
                            'confidence': 0.8,  # Default confidence for fallback
                            'reasoning': 'Fallback text parsing'
                        })
                except (ValueError, IndexError):
                    continue
        
        return results
    
    def _process_freetext_column(self, input_df: pd.DataFrame, target_df: pd.DataFrame, column_name: str, column_description: str):
        """Process columns that allow freetext input (no controlled vocabulary)"""
        
        print(f"Column definition: {column_description[:100]}..." if column_description else "No definition available")
        
        # Get sample row contexts for analysis
        row_contexts = self._extract_relevant_input_values(input_df, column_name)
        
        if not row_contexts:
            print(f"No relevant input data found for freetext column '{column_name}'")
            return
        
        # Ask AI to determine strategy for freetext column
        mapping_strategy = self._determine_freetext_strategy(row_contexts, column_name, column_description)
        
        if mapping_strategy['type'] == 'uniform':
            # Fill entire column with the same freetext value
            uniform_value = mapping_strategy['value']
            confidence = mapping_strategy['confidence']
            reasoning = mapping_strategy['reasoning']
            
            print(f"Uniform freetext for '{column_name}': '{uniform_value}' (confidence: {confidence:.2f})")
            print(f"  Reasoning: {reasoning}")
            
            # Apply uniform value to all rows
            for target_idx in target_df.index:
                target_df.at[target_idx, column_name] = uniform_value
                self.cell_mappings.append(CellMapping(
                    row_idx=target_idx,
                    column_name=column_name,
                    mapped_value=uniform_value,
                    confidence=confidence,
                    reasoning=reasoning
                ))
            
            print(f"Applied uniform freetext to {len(target_df)} rows")
            
        elif mapping_strategy['type'] == 'row_by_row':
            # Process in batches for row-by-row freetext mapping
            print(f"Row-by-row freetext mapping required for '{column_name}' - processing in batches...")
            self._apply_batch_freetext_mapping(input_df, target_df, column_name, column_description)
    
    def _determine_freetext_strategy(self, row_contexts: List[str], column_name: str, column_description: str) -> dict:
        """Determine strategy for freetext columns using structured outputs"""
        
        sample_contexts = row_contexts[:50]
        
        prompt = f"""You are an expert data curator. Determine how to GENERATE APPROPRIATE FREETEXT CONTENT for a metadata column.

TARGET COLUMN: {column_name}

COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

IMPORTANT: This is a FREETEXT column - you need to GENERATE/EXTRACT appropriate content based on the column definition, NOT map to predefined terms.

SAMPLE INPUT DATA ({len(sample_contexts)} rows):
{chr(10).join(f"Row {i+1}: {context}" for i, context in enumerate(sample_contexts))}

TASK: Determine content generation strategy based on the COLUMN DEFINITION:

1. UNIFORM - Generate the same content for all rows (rare - only if definition asks for something identical across all samples)
2. ROW_BY_ROW - Generate different content per row (typical for freetext - each row gets unique/row-specific content)

CONTENT GENERATION GUIDELINES:
- READ the column definition to understand what type of content to generate
- EXTRACT relevant information from input data that matches the definition requirements
- CREATE appropriate freetext content that fulfills the column purpose

EXAMPLES of content generation based on column definitions:
- "unique identifier" → EXTRACT: experiment_accession, sample_accession, biosample ID from each row
- "individual subject" → EXTRACT: patient ID, subject ID, or create from sample identifiers  
- "specimen identifier" → EXTRACT: sample names, specimen IDs from available metadata
- "comments/notes" → GENERATE: descriptive text about the sample/experiment from available data
- "batch information" → EXTRACT: run details, batch numbers, experiment groupings
- If definition mentions "date" → extract submission dates, collection dates
- If definition mentions "study" → extract study_title, geo_series_id, bioproject

Don't suggest generic values like "RNA-seq data" or "sequencing sample" - only specific information that the definition is asking for.

If UNIFORM: What single value from the data matches what the column definition is asking for? (Choose this ONLY if the value would truly be identical for all samples)
If ROW_BY_ROW: Explain why different rows need different values for this specific column definition. (This is the PREFERRED approach for freetext columns)"""

        try:
            # Use logged API call for freetext strategy determination
            api_response = self._make_logged_api_call(
                function_name="_determine_freetext_strategy",
                prompt=prompt,
                max_tokens=500,
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "freetext_strategy",
                        "strict": True,
                        "schema": {
                            "type": "object",
                            "properties": {
                                "type": {
                                    "type": "string",
                                    "enum": ["uniform", "row_by_row"],
                                    "description": "Whether all rows should get the same freetext value or need individual analysis"
                                },
                                "value": {
                                    "type": "string",
                                    "description": "The single freetext value if uniform, empty string if row_by_row"
                                },
                                "confidence": {
                                    "type": "number",
                                    "minimum": 0,
                                    "maximum": 1,
                                    "description": "Confidence score between 0 and 1"
                                },
                                "reasoning": {
                                    "type": "string",
                                    "description": "Brief explanation of the decision"
                                }
                            },
                            "required": ["type", "value", "confidence", "reasoning"],
                            "additionalProperties": False
                        }
                    }
                },
                temperature=0.1
            )
            
            # Parse the structured response
            strategy = json.loads(api_response['choices'][0]['message']['content'].strip())
            return {
                'type': strategy['type'],
                'value': strategy['value'],
                'confidence': float(strategy['confidence']),
                'reasoning': strategy['reasoning']
            }
                
        except Exception as e:
            print(f"Error determining freetext strategy: {e}")
            # Fallback to text parsing if structured outputs fail
            if "structured" in str(e).lower() or "json_schema" in str(e).lower():
                print("  Falling back to regular prompting for freetext strategy...")
                return self._determine_mapping_strategy_fallback(row_contexts, column_name, [], column_description)
            return {
                'type': 'row_by_row',
                'value': '',
                'confidence': 0.5,
                'reasoning': f'Error: {str(e)}'
            }
    
    def _apply_batch_freetext_mapping(self, input_df: pd.DataFrame, target_df: pd.DataFrame, column_name: str, column_description: str):
        """Apply row-by-row freetext mapping in batches"""
        
        batch_size = 20  # Reduced from 50 to avoid JSON truncation issues
        total_batches = (len(target_df) + batch_size - 1) // batch_size
        
        for batch_num in range(total_batches):
            start_idx = batch_num * batch_size
            end_idx = min(start_idx + batch_size, len(target_df))
            batch_target_rows = target_df.iloc[start_idx:end_idx]
            
            print(f"  Processing freetext batch {batch_num + 1}/{total_batches} (rows {start_idx + 1}-{end_idx})")
            
            # Use pre-computed filename mappings for this batch
            batch_mappings = self._create_batch_mappings_from_precomputed(
                input_df, batch_target_rows)
            
            if not batch_mappings:
                print(f"    No filename matches found for batch {batch_num + 1} - leaving values blank")
                continue
                
            # Ask AI to map this batch to freetext values
            batch_results = self._get_batch_freetext_mappings_with_retry(batch_mappings, column_name, column_description)
            
            # Apply the results
            for result in batch_results:
                target_idx = result['target_idx']
                value = result['value']
                confidence = result['confidence']
                reasoning = result['reasoning']
                
                if value and value.strip():  # Only apply non-empty values
                    target_df.at[target_idx, column_name] = value
                    self.cell_mappings.append(CellMapping(
                        row_idx=target_idx,
                        column_name=column_name,
                        mapped_value=value,
                        confidence=confidence,
                        reasoning=reasoning
                    ))
        
        mapped_count = sum(1 for mapping in self.cell_mappings if mapping.column_name == column_name)
        print(f"Applied {mapped_count} freetext mappings to column '{column_name}'")
    
    def _get_batch_freetext_mappings_with_retry(self, batch_mappings: List[dict], column_name: str, column_description: str) -> List[dict]:
        """Get freetext mappings with progressive retry on JSON parsing errors"""
        
        # First try with the full batch
        try:
            return self._get_batch_freetext_mappings(batch_mappings, column_name, column_description)
        except json.JSONDecodeError as e:
            print(f"    JSON parsing error for batch of {len(batch_mappings)} rows, splitting batch...")
            
            # If batch size is already 1, we can't split further
            if len(batch_mappings) <= 1:
                print(f"    Cannot split batch further, skipping row")
                return []
            
            # Split the batch in half and try each part
            mid_point = len(batch_mappings) // 2
            first_half = batch_mappings[:mid_point]
            second_half = batch_mappings[mid_point:]
            
            results = []
            
            # Try first half
            try:
                results.extend(self._get_batch_freetext_mappings_with_retry(first_half, column_name, column_description))
            except Exception as e:
                print(f"    Error processing first half of batch: {e}")
            
            # Try second half
            try:
                results.extend(self._get_batch_freetext_mappings_with_retry(second_half, column_name, column_description))
            except Exception as e:
                print(f"    Error processing second half of batch: {e}")
            
            return results
        except Exception as e:
            print(f"    Non-JSON error in batch processing: {e}")
            return []
    
    def _get_batch_freetext_mappings(self, batch_mappings: List[dict], column_name: str, column_description: str) -> List[dict]:
        """Ask AI to determine freetext values for a batch of rows"""
        
        rows_text = "\n".join([
            f"Row {i+1} (file: {mapping['filename']}): {mapping['context']}"
            for i, mapping in enumerate(batch_mappings)
        ])
        
        prompt = f"""You are an expert data curator. GENERATE APPROPRIATE FREETEXT CONTENT for multiple rows based on the column definition.

TARGET COLUMN: {column_name}

COLUMN DEFINITION: {column_description if column_description else 'No definition available'}

IMPORTANT: This is a FREETEXT column - you need to CREATE/EXTRACT appropriate content, NOT map to predefined terms.

TASK: For each row, generate freetext content that fulfills what the column definition is asking for, using the available input data.

CRITICAL INSTRUCTIONS:
1. READ the column definition carefully - it tells you EXACTLY what content to generate
2. For each row, EXTRACT or CREATE appropriate freetext that matches the definition requirements  
3. Use information from the input data to generate relevant, row-specific content
4. If the definition asks for an identifier, extract the most appropriate ID from the data
5. If the definition asks for descriptive text, create appropriate descriptions from available metadata
6. DO NOT leave fields blank - always generate appropriate content based on available data

EXAMPLES of content generation based on column definitions:
- "unique identifier" → Extract: biosample (SAMN30732287), experiment_accession (SRX17495749), sample_accession (SRS15047696)
- "individual subject" → Extract: biosample (SAMN30732287), or patient/subject identifiers from sample_title/experiment_title
- "specimen identifier" → Extract: GSM IDs (GSM6566490), sample names, or experiment identifiers  
- "parent specimen" → Extract: sample_accession (SRS15047696), biosample (SAMN30732287), or broader sample identifiers
- "comments/notes" → Extract: geo_summary, experiment_desc, study_title, or relevant descriptive text
- "batch identifier" → Extract: run_accession (SRR21492342), library information, or batch-related identifiers

4. DO NOT provide generic values like "RNA-seq data" or "sequencing sample" - extract specific identifiers, names, or descriptions
5. BE AGGRESSIVE in finding relevant content - most input data contains identifiers and information that can be extracted
6. Only leave values blank if absolutely no relevant information exists for that specific column definition

ROWS TO PROCESS:
{rows_text}

For each row, examine the data carefully and determine what freetext value should go in the '{column_name}' column.
Base your decision strictly on what the column definition is requesting."""

        try:
            # Use logged API call with structured output for freetext mappings
            api_response = self._make_logged_api_call(
                function_name="_get_batch_freetext_mappings",
                prompt=prompt,
                max_tokens=8000,  # Increased to accommodate larger batches
                response_format={
                    "type": "json_schema",
                    "json_schema": {
                        "name": "batch_freetext_mappings",
                        "strict": True,
                        "schema": {
                            "type": "object",
                            "properties": {
                                "mappings": {
                                    "type": "array",
                                    "items": {
                                        "type": "object",
                                        "properties": {
                                            "row": {
                                                "type": "integer",
                                                "minimum": 1,
                                                "description": "Row number being mapped"
                                            },
                                            "value": {
                                                "type": "string",
                                                "description": "The freetext value for this row, or empty string if no mapping"
                                            },
                                            "confidence": {
                                                "type": "number",
                                                "minimum": 0,
                                                "maximum": 1,
                                                "description": "Confidence score between 0 and 1"
                                            },
                                            "reasoning": {
                                                "type": "string",
                                                "description": "Brief explanation for this mapping"
                                            }
                                        },
                                        "required": ["row", "value", "confidence", "reasoning"],
                                        "additionalProperties": False
                                    }
                                }
                            },
                            "required": ["mappings"],
                            "additionalProperties": False
                        }
                    }
                },
                temperature=0.1
            )
            
            # Parse the structured response
            batch_result = json.loads(api_response['choices'][0]['message']['content'].strip())
            mappings = batch_result['mappings']
            
            results = []
            for mapping in mappings:
                row_num = mapping['row'] - 1  # Convert to 0-based index
                if 0 <= row_num < len(batch_mappings):
                    results.append({
                        'target_idx': batch_mappings[row_num]['target_idx'],
                        'value': mapping['value'],
                        'confidence': float(mapping['confidence']),
                        'reasoning': mapping['reasoning']
                    })
            
            return results
                
        except Exception as e:
            print(f"Error in freetext batch mapping: {e}")
            # Fallback to text parsing if structured outputs fail
            if "structured" in str(e).lower() or "json_schema" in str(e).lower():
                print("    Falling back to regular prompting for freetext batch mapping...")
                return self._get_batch_mappings_fallback(batch_mappings, column_name, [], column_description)
            return []
    
    def _map_file_format_from_row(self, input_row, column_mapping: ColumnMapping):
        """Special mapping for file format based on file extensions"""
        # Look for file_name column and extract extension
        for col in ['file_name', 'fileName', 'filename']:
            if col in input_row and pd.notna(input_row[col]):
                filename = str(input_row[col]).lower()
                
                # Map common file extensions to controlled vocabulary
                if filename.endswith('.fastq.gz') or filename.endswith('.fastq'):
                    if 'fastq' in column_mapping.controlled_vocab:
                        return 'fastq', 0.95, 'File extension indicates FASTQ format'
                elif filename.endswith('.bam'):
                    if 'bam' in column_mapping.controlled_vocab:
                        return 'bam', 0.95, 'File extension indicates BAM format'
                elif filename.endswith('.vcf'):
                    if 'vcf' in column_mapping.controlled_vocab:
                        return 'vcf', 0.95, 'File extension indicates VCF format'
                elif filename.startswith('srr') and not '.' in filename:
                    if 'sra' in column_mapping.controlled_vocab:
                        return 'sra', 0.90, 'SRR identifier indicates SRA format'
        
        return None, 0, ""
    
    def _save_excel_with_color_coding(self, df: pd.DataFrame, excel_path: str) -> None:
        """Save DataFrame to Excel with color coding based on confidence levels"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Mapped Data"
        
        # Define color schemes based on confidence levels
        colors = {
            'high': PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid'),    # Light green (0.8+)
            'medium': PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid'),  # Light yellow (0.6-0.8)
            'low': PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid'),     # Light pink (0.0-0.6)
            'original': PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid') # Lavender (pre-populated)
        }
        
        # Write data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Create a mapping of (row, column) to confidence for quick lookup
        confidence_map = {}
        for cell_mapping in self.cell_mappings:
            # Excel uses 1-based indexing, and we need to account for header row
            excel_row = cell_mapping.row_idx + 2  # +1 for 0-based to 1-based, +1 for header
            excel_col = list(df.columns).index(cell_mapping.column_name) + 1  # +1 for 1-based
            confidence_map[(excel_row, excel_col)] = cell_mapping.confidence
        
        # Apply color coding
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                # Header row - no coloring
                if row_idx == 1:
                    continue
                
                # Check if this cell was mapped by AI
                if (row_idx, col_idx) in confidence_map:
                    confidence = confidence_map[(row_idx, col_idx)]
                    if confidence >= 0.8:
                        cell.fill = colors['high']
                    elif confidence >= 0.6:
                        cell.fill = colors['medium'] 
                    else:
                        cell.fill = colors['low']
                else:
                    # Check if cell has content (pre-populated)
                    if cell.value is not None and str(cell.value).strip():
                        cell.fill = colors['original']
        
        # Add a legend worksheet
        legend_ws = wb.create_sheet("Legend")
        legend_ws['A1'] = "Color Coding Legend"
        legend_ws['A3'] = "High Confidence (0.8+)"
        legend_ws['B3'].fill = colors['high']
        legend_ws['A4'] = "Medium Confidence (0.6-0.8)"
        legend_ws['B4'].fill = colors['medium']
        legend_ws['A5'] = "Low Confidence (0.0-0.6)"
        legend_ws['B5'].fill = colors['low']
        legend_ws['A6'] = "Pre-populated Data"
        legend_ws['B6'].fill = colors['original']
        
        # Add confidence details
        if self.cell_mappings:
            legend_ws['A8'] = "AI Mapping Details"
            legend_ws['A9'] = "Column"
            legend_ws['B9'] = "Mapped Value"
            legend_ws['C9'] = "Confidence"
            legend_ws['D9'] = "Reasoning"
            
            # Group mappings by column for cleaner display
            column_mappings = {}
            for mapping in self.cell_mappings:
                if mapping.column_name not in column_mappings:
                    column_mappings[mapping.column_name] = []
                column_mappings[mapping.column_name].append(mapping)
            
            row_num = 10
            for column_name, mappings in column_mappings.items():
                # Group by unique mapped values
                unique_mappings = {}
                for mapping in mappings:
                    key = (mapping.mapped_value, mapping.confidence, mapping.reasoning)
                    if key not in unique_mappings:
                        unique_mappings[key] = 0
                    unique_mappings[key] += 1
                
                for (mapped_value, confidence, reasoning), count in unique_mappings.items():
                    legend_ws[f'A{row_num}'] = column_name
                    legend_ws[f'B{row_num}'] = mapped_value
                    legend_ws[f'C{row_num}'] = f"{confidence:.2f}"
                    legend_ws[f'D{row_num}'] = reasoning
                    if count > 1:
                        legend_ws[f'E{row_num}'] = f"({count} rows)"
                    row_num += 1
        
        # Auto-adjust column widths
        for ws_sheet in [ws, legend_ws]:
            for column in ws_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(excel_path)


def main():
    """CLI entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description="AI-powered metadata mapper")
    parser.add_argument("input_csv", help="Path to input CSV with metadata")
    parser.add_argument("target_csv", help="Path to target CSV template")
    parser.add_argument("output_csv", help="Path for output CSV with mappings")
    parser.add_argument("--config", default="config.yaml", help="Config file path")
    parser.add_argument("--creds", default="creds.yaml", help="Credentials file path")
    
    args = parser.parse_args()
    
    try:
        mapper = MetadataMapper(args.config, args.creds)
        mapper.process_csv_mapping(args.input_csv, args.target_csv, args.output_csv)
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
